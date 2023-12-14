import os
import json
import pandas as pd
from typing import Dict, Any, List, Tuple
import concurrent.futures
from schema import cv_status
from openpyxl import load_workbook
from database.db_handlers.cv_handler import CvHandler
from sqlmodel.ext.asyncio.session import AsyncSession
from service.api_service.openai_service import OpenAIService
from database.db_handlers.base_handler import BaseHandler
from database.models import CV
from fastapi import UploadFile, File, HTTPException, status
from config import TMP_PATH, JDTXT_PATH, JDSCORE_PATH, SCORING_TABLE, CVPDF_PATH, CV_EXTRACTION_PATH, JD_EXTRACTION_PATH, DUPLICATED_PATH


class DatabaseService:

    @staticmethod
    async def store_postgrescv(session: AsyncSession, cv_file: UploadFile = File(...)):
        """store cv info to postgres database"""
        try:
            if not CVPDF_PATH:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Local folder not found")
            file_cleaned = await DatabaseService.clean_filename(cv_file.filename)
            cv_data = CV(cv_status=cv_status['PROCESSING'], cv_filename=os.path.splitext(file_cleaned)[0]) # only save the filename, no extension
            print("==> Save cv to database:")
            print(cv_data)
            await CvHandler.create(session=session, model=cv_data)
            return True

        except Exception as e:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot store {cv_file.filename} detail to database | {str(e)}")


    @staticmethod
    async def store_extractjson(extracted_json, cv_file: str):
        try:
            json_file = cv_file.split(".")[0] + ".json"
            save_path = os.path.join(CV_EXTRACTION_PATH, json_file)
            with open(save_path, 'w') as file:
                json.dump(extracted_json, file)
            print(f"===> Save cv json to file: {save_path}")
            return True
        
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Can't save CV json {cv_file}")


    @staticmethod
    async def jd_store_extraction(extracted_json, jd_file: str):
        try:
            json_file = jd_file.split(".")[0] + ".json"
            save_path = os.path.join(JD_EXTRACTION_PATH, json_file)
            with open(save_path, 'w') as file:
                json.dump(extracted_json, file)
            print(f"===> Save jd json to file: {save_path}")
            return True
        
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Can't save DJ json {jd_file}")
        

    @staticmethod
    async def update_cvstatus(session: AsyncSession, cv_file: str, status: str):
#         try:
        update_data = {"cv_status": status}
        cv_filename, _ = os.path.splitext(cv_file) # store only filename in posgres db
        await CvHandler.update_by_cvname(session, cv_filename, cv_status=cv_status["READY"])
        print(" >> Update CV status")
        return True

#         except Exception:
#             raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot update cv_status {cv_file}")
        
    @staticmethod
    async def check_duplicate(file: str, path: str):
        # try:
        filename, _  = os.path.splitext(file)
        filenames = [os.path.splitext(file)[0] for file in os.listdir(path)]
        return filename in filenames

        # except Exception:
        #     raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot check duplicate {file} - {path}")
        
    
    def extract_json(json_file):
        new_data = """"""
        with open(json_file, 'r') as f:
            data_check = json.load(f)
        data = data_check["personal_information"]['description']
        new_data += "name: " + data["name"][0] + '\n'
        new_data += "earliest_university_name: " + data["earliest_university_name"][0] + '\n'
        new_data += "birthday: " + data["birthday"][0]
        return new_data      
    
    def save_xlsx(filecheck: str, results: Any, idx: int, wb):
        sh = wb.active
        sim_files = []
        for filename, result in results:
            sh[f'A{idx}'] = filename[:-5]+'.pdf'
            sh[f'B{idx}'] = result['score']
            sh[f'C{idx}'] = result['explanation']
            if result['score'] == 10:
                sim_files.append(filename[:-5]+'.pdf')
            idx += 1   
        wb.save(os.path.join(DUPLICATED_PATH, f"{filecheck}.xlsx"))
        print(f" >> Save results in {DUPLICATED_PATH}")
        return sim_files
    
    def fill_info():
        file_info = ()
        for file in os.listdir(DUPLICATED_PATH):
            wb = load_workbook(os.path.join(DUPLICATED_PATH, file))
            sheet = wb.worksheets[0]
            info = (file, sheet.max_row-1)
            file_info += (info,)
        
        for curr_name, curr_row in file_info:
            #   Load current file
            wb = load_workbook(os.path.join(DUPLICATED_PATH, curr_name))
            sh = wb.active
            for filename, num_row in file_info:
                if curr_row < num_row:
                    #   Convert xlsx to csv for easy works (Detect missing rows)
                    df = pd.DataFrame(pd.read_excel(os.path.join(DUPLICATED_PATH, filename)))
                    sub_df = df[df['filename'] == f'{curr_name[:-5]}.pdf']
                    data = list(sub_df.values[0])
                    data[0] = filename[:-5]+'.pdf'
                    #   Append rows to missing xlsx file
                    sh.append(data)
            wb.save(os.path.join(DUPLICATED_PATH, curr_name))
                    
        
    @staticmethod
    async def check_dup(diffs: str):        
        template = """
        Please compare and score the similarity of the [File 1] to [File 2] provided below on a scale of 10 and give a detailed explanation. The higher the score, the more similar the two Files are.
        [File 1]
        {file1}

        [File 2]
        {file2}

        The reponse must follow strictly the below JSON format:
        {{
            score:
            explanation:
        }}
        """        
        
        for filename in diffs:
            #   Load uploaded json file that needs to check duplicate
            cv_check = DatabaseService.extract_json(os.path.join(CV_EXTRACTION_PATH, filename+'.json'))
            
            cv_names = []
            prompts = []
            #   Load exracted json file
            for file in os.listdir(CV_EXTRACTION_PATH):
                if file.split('.')[0] != filename:
                    cv_extracted = DatabaseService.extract_json(os.path.join(CV_EXTRACTION_PATH, file))
                    prompt = template.format(file1=cv_check, file2=cv_extracted)
                    prompts.append(prompt)
                    cv_names.append(file)   
            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                # Calling openai
                results = list(executor.map(OpenAIService.gpt_api, prompts, cv_names))  
                
            #   Save to xlsx file  
            wb = load_workbook("resources/template/check_dup.xlsx")
            idx = 2
            sim_files = DatabaseService.save_xlsx(filename, results, idx, wb)  
            if sim_files:
                print(f" >> Detected duplicate file(s) with {filename}.pdf:", ', '.join(sim_files)) 
            else:
                print(f" >> Congrats! You don't have any duplicate file with {filename}.pdf\n")
        

    @staticmethod
    async def create_scoreparquet():
        """`check if jd existed yet, if not, create a score parquet file with initial column created`"""
        try:
            for jd_file in os.listdir(JDTXT_PATH):
                if not await DatabaseService.check_duplicate(jd_file, JDSCORE_PATH):
                    await DatabaseService.create_emptyparquet(jd_file, SCORING_TABLE)
            return True
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Can't create score parquet")
    
    
    @staticmethod
    async def create_emptyparquet(jd_file: str, columns_dict: dict):
        """create empty parquet file store blank table"""
        try:
            data = pd.DataFrame(columns_dict)
            jd_filename = os.path.splitext(jd_file)[0]
            data.to_parquet(f"{JDSCORE_PATH}/{jd_filename}.parquet")
            return True
        
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Can't create empty score parquet")
        
    @staticmethod
    async def check_duplicatescore(cv_file: str, jd_file: str):
        """check if the cv filename is existed in the score table of jd"""
        try:
            cv_name, _ = os.path.splitext(cv_file)
            jd_dir = [os.path.splitext(file)[0] for file in os.listdir(os.path.join(JDSCORE_PATH, jd_file[:-5]))]
            return cv_name in jd_dir
        
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot check cv score duplicate")
        
    @staticmethod
    async def add_cvscore(jd_filename: str, cv_filename: str, score_dict: Dict):
        """add cv score to jd score local folder"""
        # try:
        if not os.path.exists(os.path.join(JDSCORE_PATH, jd_filename, f"{jd_filename}.xlsx")):
            wb = load_workbook("resources/template/scoring.xlsx")
        else:
            wb = load_workbook(os.path.join(JDSCORE_PATH, jd_filename, f"{jd_filename}.xlsx"))
        sh = wb.active
        idx = sh.max_row + 1
        
        data_dict = {"B": "job_title", "D": "job_summary", "F": "qualifications_and_skills", "H": "education", "J": "knowledge_and_experience", "L": "benefits_and_salary", "N": "overall"}
        sh[f"A{idx}"] = cv_filename+'.pdf'
        for key, val in data_dict.items():
            if val == "overall":
                sh[key+str(idx)] = score_dict[val]["overall_score"]
                key_next = chr(ord(key) + 1)
                sh[key_next+str(idx)] = score_dict[val]["overall_evaluation"]
            else:
                sh[key+str(idx)] = score_dict[val]["score"]
                key_next = chr(ord(key) + 1)
                sh[key_next+str(idx)] = score_dict[val]["explanation"]
                
        #   Save file.xlsx output
        wb.save(os.path.join(JDSCORE_PATH, jd_filename, f"{jd_filename}.xlsx"))
        
        return True
        
        # except Exception:
        #     raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Can't save cvscore to parquet.")

    @staticmethod
    async def save_cvpdf(cv_file: UploadFile = File(...)):
        """save cv pdf to local folder"""
        try:
            if not os.path.exists(CVPDF_PATH):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Folder not exist {CVPDF_PATH}")
            contents = cv_file.file.read()
            cv_file_cleaned = await DatabaseService.clean_filename(cv_file.filename)
            with open(f"{CVPDF_PATH}/{cv_file_cleaned}", 'wb') as f:
                f.write(contents)
            print(f"==> Save pdf file to path: {CVPDF_PATH}/{cv_file_cleaned}")
            return True
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot read file {cv_file_cleaned}")

    @staticmethod
    async def save_jdtxt(jd_file: UploadFile = File(...)):
        """save jd text file to local folder"""
        try:
            if not os.path.exists(JDTXT_PATH):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Folder not exist {CVPDF_PATH}")
            contents = jd_file.file.read()
            jd_file_cleaned = await DatabaseService.clean_filename(jd_file.filename)
            #   Read and save as pdf file
            with open(f"{JDTXT_PATH}/{jd_file_cleaned}", 'wb') as f:
                f.write(contents)
            print(f"==> Save {jd_file_cleaned} to {JDTXT_PATH}/{os.path.splitext(jd_file_cleaned)[0]}.txt")
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot save file {jd_file_cleaned}")
            

    @staticmethod
    async def clean_filename(filename):
        # Split the filename and file extension
        name, extension = os.path.splitext(filename)
        # Replace spaces with underscores in the filename
        name = name.replace(" ", "_")
        # Remove dots from the filename
        name = name.replace(".", "")
        # Concatenate the cleaned filename with the file extension
        cleaned_filename = name + extension
        return cleaned_filename
